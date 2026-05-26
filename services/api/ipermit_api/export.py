"""Deliverable exports for the permit matrix (S02-04, ADR-0003 contract 4, T07-04).

Renders a persisted permit matrix to ``json`` / ``xlsx`` / ``pdf``. The export
layer renders the *same* tier/citation/advisory data the matrix already carries —
it never recomputes or strips it (ADR-0003). Every deliverable includes the
standing platform disclaimer (T00-05 §4.4) and each permit's advisories, so no
export can omit the liability language (AGENTS.md *Liability Strategy*).
"""

from __future__ import annotations

import io
import json
from typing import Any

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .envelope import PLATFORM_ADVISORY

#: Supported deliverable formats.
EXPORT_FORMATS = ("json", "xlsx", "pdf")

_MEDIA_TYPES = {
    "json": "application/json",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}

_PERMIT_COLUMNS = (
    "permit_name",
    "agency",
    "jurisdiction_level",
    "tier",
    "tier_rationale",
    "citations",
    "advisories",
    "known_unknowns",
    "stage",
)


def export_matrix(matrix: dict[str, Any], fmt: str) -> tuple[bytes, str, str]:
    """Render *matrix* to *fmt*; return (content, media_type, filename)."""
    if fmt not in EXPORT_FORMATS:
        raise ValueError(f"unsupported export format {fmt!r}")
    renderer = {"json": _to_json, "xlsx": _to_xlsx, "pdf": _to_pdf}[fmt]
    content = renderer(matrix)
    filename = f"permit-matrix-{matrix['evaluation_id']}.{fmt}"
    return content, _MEDIA_TYPES[fmt], filename


def _to_json(matrix: dict[str, Any]) -> bytes:
    """Deliverable = the matrix plus the standing platform advisory."""
    payload = {"matrix": matrix, "advisories": [PLATFORM_ADVISORY]}
    return json.dumps(payload, indent=2, ensure_ascii=True).encode("utf-8")


def _permit_row(permit: dict[str, Any]) -> list[str]:
    """Flatten a permit into the ordered string cells of the table."""
    return [
        permit["permit_name"],
        permit["agency"],
        permit["jurisdiction_level"],
        str(permit["confidence"]["tier"]),
        permit["confidence"]["tier_rationale"],
        "\n".join(c["reference"] for c in permit["citations"]),
        "\n".join(a["text"] for a in permit["advisories"]),
        "\n".join(k["text"] for k in permit["known_unknowns"]),
        str(permit["sequencing"]["stage_index"]),
    ]


def _to_xlsx(matrix: dict[str, Any]) -> bytes:
    """One sheet: a metadata/advisory header block, then the permit table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Permit Matrix"
    ws.append(["iPermit — Permit Matrix (advisory only)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Evaluation", matrix["evaluation_id"]])
    ws.append(["Evaluation date", matrix["evaluation_date"]])
    ws.append(["Ruleset", matrix["ruleset_version"]["content_hash"]])
    ws.append(["Advisory", PLATFORM_ADVISORY["text"]])
    ws.append([])
    ws.append([c.replace("_", " ").title() for c in _PERMIT_COLUMNS])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for permit in matrix["permits"]:
        ws.append(_permit_row(permit))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 28
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _line(pdf: FPDF, height: float, text: str) -> None:
    """Write one wrapped line, returning the cursor to the left margin."""
    pdf.multi_cell(0, height, text, new_x=XPos.LMARGIN, new_y=YPos.NEXT)


def _pdf_permit_block(pdf: FPDF, permit: dict[str, Any]) -> None:
    """Write one permit's section to the PDF."""
    pdf.set_font("Helvetica", "B", 11)
    _line(pdf, 6, f"[Tier {permit['confidence']['tier']}] {permit['permit_name']}")
    pdf.set_font("Helvetica", "", 9)
    _line(pdf, 5, f"Agency: {permit['agency']} ({permit['jurisdiction_level']})")
    _line(pdf, 5, permit["confidence"]["tier_rationale"])
    for citation in permit["citations"]:
        _line(pdf, 5, f"  - {citation['reference']}")
    for advisory in permit["advisories"]:
        _line(pdf, 5, f"  ! {advisory['text']}")
    pdf.ln(2)


def _to_pdf(matrix: dict[str, Any]) -> bytes:
    """A simple linear PDF: title, metadata, platform advisory, per-permit blocks."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    _line(pdf, 8, "iPermit - Permit Matrix")
    pdf.set_font("Helvetica", "", 9)
    _line(pdf, 5, f"Evaluation {matrix['evaluation_id']}")
    _line(pdf, 5, f"Evaluation date: {matrix['evaluation_date']}")
    pdf.set_font("Helvetica", "I", 9)
    _line(pdf, 5, PLATFORM_ADVISORY["text"])
    pdf.ln(3)
    for permit in matrix["permits"]:
        _pdf_permit_block(pdf, permit)
    return bytes(pdf.output())
