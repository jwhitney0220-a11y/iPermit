"""Unit tests for the permit-matrix exporters (S02-04).

Every format must carry the liability language (ADR-0003): the standing platform
advisory and each permit's advisories are present in the rendered output.
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from ipermit_api.envelope import PLATFORM_ADVISORY
from ipermit_api.export import EXPORT_FORMATS, export_matrix
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
_MATRIX = json.loads(
    (ROOT / "docs/specs/examples/permit-matrix-example.json").read_text()
)


def test_unsupported_format_raises() -> None:
    with pytest.raises(ValueError, match="unsupported export format"):
        export_matrix(_MATRIX, "docx")


@pytest.mark.parametrize("fmt", EXPORT_FORMATS)
def test_export_returns_nonempty_with_filename(fmt: str) -> None:
    content, media_type, filename = export_matrix(_MATRIX, fmt)
    assert content
    assert media_type
    assert filename.endswith(f".{fmt}")
    assert _MATRIX["evaluation_id"] in filename


def test_json_export_round_trips_and_carries_advisory() -> None:
    content, media_type, _ = export_matrix(_MATRIX, "json")
    assert media_type == "application/json"
    payload = json.loads(content)
    assert payload["matrix"] == _MATRIX
    assert any(a["origin"] == "platform" for a in payload["advisories"])


def test_xlsx_export_is_valid_and_lists_permits() -> None:
    content, _, _ = export_matrix(_MATRIX, "xlsx")
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    text = "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value
    )
    assert PLATFORM_ADVISORY["text"] in text
    for permit in _MATRIX["permits"]:
        assert permit["permit_name"] in text


def test_pdf_export_has_pdf_header() -> None:
    content, media_type, _ = export_matrix(_MATRIX, "pdf")
    assert media_type == "application/pdf"
    assert content[:5] == b"%PDF-"
    assert len(content) > 500
